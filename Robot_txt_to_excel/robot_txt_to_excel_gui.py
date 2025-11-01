#Esse código se trata de um robô com interface gráfica (GUI) usando Tkinter e OpenPyXL, que permite:
# Escolher o arquivo .txt
# Gerar automaticamente o .xlsx com colunas: Nome, Telefone, Endereço, extraindo as informações do txt.
# Exibir mensagens de sucesso ou erro na tela.
# Para rodar instale o pacode com o comando: pip install openpyxl
# E execute com o comando: python robot_txt_to_excel.py

import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox

def converter_txt_para_excel():
    try:
        # Seleciona o arquivo TXT
        arquivo_txt = filedialog.askopenfilename(
            title="Selecione o arquivo TXT",
            filetypes=[("Arquivos de texto", "*.txt")]
        )

        if not arquivo_txt:
            return  # usuário cancelou

        # Escolhe onde salvar o Excel
        arquivo_excel = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel como",
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )

        if not arquivo_excel:
            return  # usuário cancelou

        # Cria uma nova planilha
        wb = openpyxl.Workbook()
        planilha = wb.active
        planilha.title = "Contatos"

        # Cabeçalhos
        cabecalhos = ["Nome", "Telefone", "Endereço"]
        for col, cab in enumerate(cabecalhos, start=1):
            planilha.cell(row=1, column=col, value=cab)

        # Lê o arquivo TXT
        with open(arquivo_txt, "r", encoding="utf-8") as f:
            for i, linha in enumerate(f, start=2):
                valores = [v.strip() for v in linha.strip().split(",")]
                for j, valor in enumerate(valores, start=1):
                    planilha.cell(row=i, column=j, value=valor)

        # Ajusta largura das colunas
        planilha.column_dimensions["A"].width = 30
        planilha.column_dimensions["B"].width = 18
        planilha.column_dimensions["C"].width = 50

        # Salva o Excel
        wb.save(arquivo_excel)

        messagebox.showinfo("Sucesso", f"Arquivo Excel criado com sucesso:\n{arquivo_excel}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")

# --- Interface gráfica ---
janela = tk.Tk()
janela.title("Conversor TXT → Excel")
janela.geometry("400x200")
janela.resizable(False, False)

titulo = tk.Label(janela, text="Conversor de Lista TXT para Excel", font=("Arial", 14, "bold"))
titulo.pack(pady=20)

botao_converter = tk.Button(
    janela,
    text="Selecionar arquivo TXT e converter",
    command=converter_txt_para_excel,
    font=("Arial", 12),
    bg="#4CAF50",
    fg="white",
    padx=10,
    pady=5
)
botao_converter.pack(pady=20)

rodape = tk.Label(janela, text="Desenvolvido em Python + Tkinter", font=("Arial", 9), fg="gray")
rodape.pack(side="bottom", pady=10)

janela.mainloop()
